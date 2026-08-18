#!/usr/bin/env python3
"""Build the web app's data bundle from an ICL master extraction workbook."""

from __future__ import annotations

import argparse
import json
import math
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


PATTERN_LABELS = {
    "monotonic_improvement": "Monotonic improvement",
    "monotonic_decline": "Monotonic decline",
    "improve_then_decline": "Improve then decline",
    "decline_then_recover": "Decline then recover",
    "mixed_nonmonotonic": "Mixed / non-monotonic",
    "flat_or_stable": "Flat / stable",
    "two_point_improvement": "Two-point improvement",
    "two_point_decline": "Two-point decline",
    "insufficient_points": "Insufficient points",
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def rows_for(sheet):
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "" for value in next(rows)]
    return [
        {header: clean(value) for header, value in zip(headers, row) if header}
        for row in rows
        if any(value is not None for value in row)
    ]


def metric_group(metric):
    text = str(metric or "").lower()
    if "cat attentiveness" in text:
        return "Attentiveness"
    if "accuracy" in text or text.startswith("acc@"):
        return "Accuracy"
    if "entity-level f1" in text or text == "f1":
        return "F1"
    if any(token in text for token in ("spbleu", "chrf", "comet", "bleurt")):
        return "Machine translation quality"
    if "rouge" in text or "bertscore" in text:
        return "Text generation similarity"
    if any(
        token in text
        for token in (
            "ca-score",
            "clipscore",
            "longest common",
            "kendall",
            "distance",
            "minimum swap",
        )
    ):
        return "Procedural ordering / similarity"
    return "Other"


def task_group(task):
    text = str(task or "").lower()
    if "translation" in text or "post-editing" in text:
        return "Machine translation"
    if "named entity recognition" in text:
        return "Information extraction"
    if "question answering" in text:
        return "Question answering"
    if any(token in text for token in ("natural language inference", "entailment", "fact verification")):
        return "Inference and verification"
    if "paraphrase" in text:
        return "Paraphrase detection"
    if "visual step" in text or "cross-modal step" in text:
        return "Multimodal procedural reasoning"
    if "explanation generation" in text:
        return "Explanation generation"
    if any(token in text for token in ("classification", "sentiment", "topic")):
        return "Classification"
    return "Other"


def evidence_tier(number_of_conditions):
    count = int(number_of_conditions or 0)
    if count >= 3:
        return "3+ points"
    if count == 2:
        return "2 points"
    return "Insufficient"


def pick(record, *keys):
    return {key: record.get(source) for key, source in keys}


def build_bundle(workbook_path: Path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    trajectories = rows_for(workbook["Trajectories"])
    results = rows_for(workbook["All Results"])
    tests = rows_for(workbook["Statistical Tests"])

    results_by_trajectory = {}
    for row in results:
        trajectory_id = str(row.get("trajectory_id") or "")
        if not trajectory_id:
            continue
        point = pick(
            row,
            ("shotCount", "shot_count"),
            ("rawScoreReported", "raw_score_reported"),
            ("rawScore", "raw_score_numeric"),
            ("analysisScore", "analysis_score"),
            ("changeFromBaseline", "change_from_baseline"),
            ("percentChangeFromBaseline", "percent_change_from_baseline"),
            ("stepChange", "step_change"),
            ("standardDeviation", "standard_deviation"),
            ("standardError", "standard_error"),
            ("confidenceIntervalLower", "confidence_interval_lower"),
            ("confidenceIntervalUpper", "confidence_interval_upper"),
            ("confidenceLevel", "confidence_level"),
        )
        results_by_trajectory.setdefault(trajectory_id, []).append(point)

    for points in results_by_trajectory.values():
        points.sort(key=lambda point: (point.get("shotCount") is None, point.get("shotCount") or 0))

    tests_by_trajectory = {}
    for row in tests:
        trajectory_id = str(row.get("trajectory_id") or "")
        if not trajectory_id:
            continue
        test = pick(
            row,
            ("baselineShot", "comparison_baseline_shot"),
            ("targetShot", "comparison_target_shot"),
            ("pValue", "p_value"),
            ("adjustedPValue", "adjusted_p_value"),
            ("posteriorProbability", "posterior_probability"),
            ("statisticallySignificant", "statistically_significant"),
            ("effectSizeName", "effect_size_name"),
            ("effectSizeValue", "effect_size_value"),
            ("confidenceIntervalLower", "confidence_interval_lower"),
            ("confidenceIntervalUpper", "confidence_interval_upper"),
            ("confidenceLevel", "confidence_level"),
            ("testName", "statistical_test_name"),
            ("testType", "statistical_test_type"),
            ("sidedness", "test_sidedness"),
            ("significanceLevel", "significance_level"),
            ("multipleComparisonCorrection", "multiple_comparison_correction"),
            ("correctionMethod", "correction_method"),
            ("description", "statistical_test_description"),
            ("interpretation", "statistical_test_interpretation"),
            ("table", "statistical_table"),
            ("tableLink", "statistical_table_link"),
            ("notes", "notes"),
        )
        tests_by_trajectory.setdefault(trajectory_id, []).append(test)

    web_trajectories = []
    for row in trajectories:
        trajectory_id = str(row.get("trajectory_id") or "")
        pattern = str(row.get("numerical_pattern") or "insufficient_points")
        normalized_change = row.get("normalized_endpoint_change")
        if normalized_change is None:
            endpoint_outcome = "Unknown"
        elif normalized_change > 0:
            endpoint_outcome = "Improved"
        elif normalized_change < 0:
            endpoint_outcome = "Declined"
        else:
            endpoint_outcome = "No change"

        web_trajectories.append(
            {
                "trajectoryId": trajectory_id,
                "paperId": row.get("paper_id"),
                "paperTitle": row.get("paper_title"),
                "task": row.get("task"),
                "taskGroup": task_group(row.get("task")),
                "dataset": row.get("dataset"),
                "modelName": row.get("model_name"),
                "modelType": row.get("model_type"),
                "metric": row.get("metric"),
                "metricGroup": metric_group(row.get("metric")),
                "metricDirection": row.get("metric_direction"),
                "metricDescription": row.get("metric_description"),
                "metricScale": row.get("metric_scale_or_range"),
                "metricInterpretation": row.get("metric_interpretation"),
                "numberOfConditions": row.get("number_of_conditions"),
                "lowestShot": row.get("lowest_tested_shot"),
                "highestShot": row.get("highest_tested_shot"),
                "baselineRawScore": row.get("baseline_raw_score"),
                "endpointRawScore": row.get("endpoint_raw_score"),
                "normalizedEndpointChange": normalized_change,
                "endpointOutcome": endpoint_outcome,
                "bestShotCount": row.get("best_shot_count"),
                "maxDrawdown": row.get("max_drawdown"),
                "pattern": pattern,
                "category": PATTERN_LABELS.get(pattern, pattern.replace("_", " ").title()),
                "evidenceTier": evidence_tier(row.get("number_of_conditions")),
                "numericalDegradation": bool(row.get("numerical_degradation")),
                "statisticallySupportedDegradation": bool(row.get("statistically_supported_degradation")),
                "statisticalTestUsed": row.get("statistical_test_used"),
                "resultTable": row.get("result_table"),
                "resultTableLink": row.get("result_table_link"),
                "verificationStatus": row.get("verification_status"),
                "notes": row.get("notes"),
                "results": results_by_trajectory.get(trajectory_id, []),
                "statisticalTests": tests_by_trajectory.get(trajectory_id, []),
            }
        )

    paper_ids = {row["paperId"] for row in web_trajectories if row.get("paperId")}
    return {
        "meta": {
            "schemaVersion": 1,
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "sourceFile": workbook_path.name,
            "paperCount": len(paper_ids),
            "trajectoryCount": len(web_trajectories),
            "resultCount": sum(len(row["results"]) for row in web_trajectories),
            "verificationNotice": "All current trajectories require source-table verification before publication.",
        },
        "trajectories": web_trajectories,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path, help="Path to the master extraction workbook")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("app/data/atlas.json"),
        help="Destination JSON file",
    )
    args = parser.parse_args()
    bundle = build_bundle(args.workbook.resolve())
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(bundle, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(
        f"Wrote {args.output}: {bundle['meta']['paperCount']} papers, "
        f"{bundle['meta']['trajectoryCount']} trajectories, {bundle['meta']['resultCount']} shot results"
    )


if __name__ == "__main__":
    main()

